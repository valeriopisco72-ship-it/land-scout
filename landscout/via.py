"""land-scout — modulo VIA (v0.1, 15/07/2026): censimento NAZIONALE dei progetti dal portale VIA.

Sblocca la geografia: prima il censimento (139 progetti BN/AV) era compilato a mano,
quindi `match.py` funzionava solo nel Sannio. Qui si costruisce da solo, per tutta Italia.

Come funziona (scoperto ispezionando il portale):
  - la ricerca di va.mite.gov.it e' un form **GET** (niente token/POST)
  - esiste un **export XLSX**: ?Testo=<termine>&t=o&mode=export   <- il parametro `t=o` e' obbligatorio
    (senza, risponde 404)
  - l'export ha 3 colonne: Progetto | Proponente | Ultima procedura
  - NON contiene provincia/MW/link: provincia, comuni e potenza si estraggono dal testo
    del progetto (resa: prov 91%, MW 94%, comuni 95% su 865 righe agrivoltaico)

Il link al fascicolo VIA non e' nell'export: le righe curate a mano (BN/AV) lo hanno e
vengono preservate nel merge.

CLI:
  .venv/Scripts/python -m landscout.via --build          # ricostruisce il censimento nazionale
  .venv/Scripts/python -m landscout.via --build --termini agrivoltaico,eolico
"""
import argparse, csv, re, sys, warnings
from pathlib import Path
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
try: sys.stdout.reconfigure(encoding='utf-8')
except Exception: pass
warnings.filterwarnings('ignore')
from landscout.config import EP, CENSUS, TTL_GIORNI
from landscout.cache import cached_file

# termine di ricerca -> categoria (match.norm_tech sa mapparle in agriPV/PV/BESS/wind)
TERMINI = {
    'agrivoltaico': 'AGRIVOLTAICO',
    'fotovoltaico': 'FOTOVOLTAICO',
    'eolico':       'EOLICO',
    'accumulo':     'ACCUMULO',
}
# priorita' categoria in caso di doppione (un agrivoltaico contiene anche "fotovoltaico")
PRIORITA = {'AGRIVOLTAICO': 0, 'ACCUMULO': 1, 'EOLICO': 2, 'FOTOVOLTAICO': 3}

RE_PROV = re.compile(r'\(([A-Z]{2})\)')
RE_MW = re.compile(r'(\d+(?:[.,]\d+)?)\s*MW', re.I)
RE_COM = re.compile(r'[Cc]omun[ei]\s+di\s+([^().;]+)')
CAMPI = ['categoria', 'prov', 'proponente', 'MW', 'progetto', 'procedura', 'link']


def fetch_export(termine, forza=False):
    """Scarica l'export XLSX per un termine. Ritorna lista di (progetto, proponente, procedura)."""
    import openpyxl
    url = f"{EP['via_export']}?Testo={termine}&t=o&mode=export"
    p = cached_file(url, f'via_{termine}.xlsx', ttl_giorni=TTL_GIORNI.get('census', 30), forza=forza)
    if p is None:
        print(f'  ! export "{termine}" non scaricabile'); return []
    try:
        wb = openpyxl.load_workbook(p)          # NB: read_only sbaglia le dimensioni (manca <dimension>)
        rows = list(wb[wb.sheetnames[0]].iter_rows(values_only=True))
    except Exception as e:
        print(f'  ! export "{termine}" illeggibile: {e}'); return []
    return [r for r in rows[1:] if r and r[0]]


def parse_riga(prog, prop, proc, categoria):
    t = prog or ''
    prov = RE_PROV.findall(t)
    mws = [float(m.replace('.', '').replace(',', '.')) for m in RE_MW.findall(t)]
    comuni = []
    for m in RE_COM.findall(t):
        for part in re.split(r',|\se\s', m):
            n = part.strip().strip('"')
            if n and len(n) < 40 and n[:1].isupper():
                comuni.append(n)
    return {'categoria': categoria,
            'prov': prov[0] if prov else '',
            'proponente': (prop or '').strip(),
            'MW': f'{max(mws):.3f}' if mws else '',          # la potenza dell'impianto = la piu' grande citata
            'progetto': ' '.join(t.split()),
            'procedura': (proc or '').strip(),
            'link': ''}


def build(termini=None, forza=False):
    termini = termini or list(TERMINI)
    trovati = {}
    for t in termini:
        cat = TERMINI.get(t, t.upper())
        rows = fetch_export(t, forza=forza)
        print(f'  {t:14} -> {len(rows):>4} progetti')
        for prog, prop, proc in rows:
            r = parse_riga(prog, prop, proc, cat)
            key = (r['proponente'].lower(), r['progetto'][:80].lower())
            old = trovati.get(key)
            if old is None or PRIORITA.get(cat, 9) < PRIORITA.get(old['categoria'], 9):
                trovati[key] = r
    # --- merge con il censimento curato a mano (ha i LINK, che l'export non da') ---
    curati = 0
    if Path(CENSUS).exists():
        with open(CENSUS, encoding='utf-8-sig') as f:
            for c in csv.DictReader(f, delimiter=';'):
                if not (c.get('proponente') or '').strip():
                    continue
                key = ((c['proponente'] or '').strip().lower(), ' '.join((c.get('progetto') or '').split())[:80].lower())
                if key in trovati:
                    for campo in ('link', 'prov', 'MW'):        # il curato vince: e' verificato
                        if c.get(campo):
                            trovati[key][campo] = c[campo]
                else:
                    trovati[key] = {k: (c.get(k) or '') for k in CAMPI}
                curati += 1
    print(f'  merge con censimento curato: {curati} righe (link preservati)')
    return sorted(trovati.values(), key=lambda r: (r['categoria'], r['prov'], r['proponente']))


def scrivi(righe, dest=None):
    dest = Path(dest or CENSUS)
    dest.parent.mkdir(parents=True, exist_ok=True)
    with open(dest, 'w', newline='', encoding='utf-8-sig') as f:
        w = csv.DictWriter(f, fieldnames=CAMPI, delimiter=';')
        w.writeheader()
        for r in righe:
            w.writerow(r)
    return dest


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--build', action='store_true')
    ap.add_argument('--termini', help='lista separata da virgola (default: tutti)')
    ap.add_argument('--forza', action='store_true', help='ignora la cache e riscarica')
    ap.add_argument('--out')
    A = ap.parse_args()
    if not A.build:
        print('Uso: --build [--termini agrivoltaico,eolico] [--forza]'); return
    termini = [t.strip() for t in A.termini.split(',')] if A.termini else None
    print('Censimento VIA nazionale — scarico gli export...')
    righe = build(termini, forza=A.forza)
    dest = scrivi(righe, A.out)
    from collections import Counter
    cat = Counter(r['categoria'] for r in righe)
    prov = Counter(r['prov'] for r in righe if r['prov'])
    print(f'\nTOTALE: {len(righe)} progetti -> {dest}')
    print('  per categoria:', dict(cat))
    print(f'  province coperte: {len(prov)}  | top: {dict(prov.most_common(8))}')
    print(f'  con link: {sum(1 for r in righe if r["link"])} | con prov: {sum(1 for r in righe if r["prov"])}')


if __name__ == '__main__':
    main()
