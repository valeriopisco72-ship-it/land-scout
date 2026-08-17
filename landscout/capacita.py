"""land-scout capacita' — il nodo di rete regge, o e' gia' pieno?

E' il gate che puo' azzerare tutto il resto: un blocco perfetto su una rete
satura non si connette, e mesi di aggregazione fondiaria non servono a niente.
Va guardato PRIMA di spendere su visure e trattative, non dopo.

DUE DOMANDE DIVERSE, DUE FONTI DIVERSE
--------------------------------------
1. **La rete di distribuzione e' congestionata?** → mappa "Aree critiche" di
   e-Distribuzione, aggiornata mensilmente sulle sezioni AT/MT delle cabine
   primarie. Il feature service ArcGIS che alimenta la mappa pubblica e'
   interrogabile: `criticita_provincia()`. **Granularita' PROVINCIALE**: dice
   se la provincia e' messa male, non se quella specifica cabina lo e'.
2. **Quanta capacita' e' gia' stata chiesta da altri?** → la coda di
   connessione (Terna Econnextion). Questa **non e' interrogabile**: e' una
   dashboard PowerBI. Si esporta a mano e `coda_da_export()` la legge — stesso
   patto di `visure.py`: l'export lo fai tu, la lettura la fa il tool.

SULLA SCALA DI CRITICITA'
-------------------------
Il servizio espone `livellocriticita` come intero 0-4 **senza documentazione
pubblica del significato**. L'ordinamento e' stato ricavato dai dati stessi
confrontando province a saturazione nota: Milano 1 · Napoli 2 · Benevento e
Avellino 3 · **Foggia e Lecce 4** — cioe' le due province piu' sature d'Italia
stanno in cima. Crescente = peggio. E' una lettura per evidenza, non una fonte
ufficiale: il modulo la dichiara come tale e non la spaccia per certa.

Il campo `livellocriticita2` (che la mappa usa per il trimestre successivo) vale
1 per 107 province su 110: non e' una previsione utilizzabile e **non viene
usata**. Meglio un dato in meno che un dato finto.

Uso:
    from landscout import capacita
    c = capacita.criticita_provincia('Benevento')
    q = capacita.coda_da_export('Terna_Econnextion_....xlsx', prov='BENEVENTO')
    print(capacita.print_valutazione(capacita.valuta(c, q, mwp=(5.6, 8.9))))
"""
import json
import os
import re
import urllib.parse
import urllib.request

FS = ('https://dpa-portalgis.enel.com/server/rest/services/Hosted/'
      'ProvinceCritiche_View/FeatureServer/0')
UA = {'User-Agent': 'land-scout capacita', 'Accept': 'application/json'}

# Lettura per EVIDENZA (vedi docstring), non da documentazione ufficiale.
SCALA = {
    0: ('dato assente', 'la provincia non risulta nel servizio (spesso: non concessionaria)'),
    1: ('bassa', 'rete con margine: nessun segnale di congestione'),
    2: ('media', 'primi segnali di saturazione'),
    3: ('alta', 'rete congestionata: connessione onerosa o con attese lunghe'),
    4: ('molto alta', 'fra le province piu' + "'" + ' sature: connessione il rischio n.1'),
}
# Stati della coda Terna, dal piu' vincolante al meno.
PESO_STATO = {
    'in esercizio': 1.0,
    'autorizzati': 0.9,
    'progetti con nulla osta': 0.8,
    'stmg accettate': 0.6,
    'stmg emesse': 0.35,
    'richieste': 0.15,
}


def _get(url, timeout=60):
    return json.loads(urllib.request.urlopen(
        urllib.request.Request(url, headers=UA), timeout=timeout).read())


def criticita_provincia(nome=None, cod_pro=None, cache_dir=None, timeout=60):
    """Livello di criticita' della rete di distribuzione per provincia.

    Fonte: feature service pubblico che alimenta la mappa "Aree critiche" di
    e-Distribuzione. Se la rete non risponde si restituisce `livello=None` e
    `verificato=False`: mai un "nessuna criticita'" inventato.
    """
    where = (f"UPPER(f4)=UPPER('{nome}')" if nome else f"cod_pro={int(cod_pro)}")
    q = FS + '/query?' + urllib.parse.urlencode({
        'where': where, 'outFields': 'f2,f4,cod_pro,livellocriticita',
        'returnGeometry': 'false', 'f': 'json'})
    try:
        d = _get(q, timeout)
        # ArcGIS risponde 200 anche quando il servizio non c'e' piu': l'errore sta
        # nel corpo, non nello stato HTTP. Senza questo controllo un servizio
        # spostato diventa "provincia non presente" — cioe' un guasto travestito da
        # dato verificato. (Trovato in audit il 08/08/2026 interrogando una URL
        # sbagliata: la risposta era {"error":{"code":404,...}} con HTTP 200.)
        if d.get('error'):
            e = d['error']
            return {'provincia': nome, 'livello': None, 'verificato': False,
                    'nota': f"servizio in errore ({e.get('code')}: {e.get('message')}): "
                            f'criticita NON verificata'}
        f = (d.get('features') or [])
        if not f:
            return {'provincia': nome, 'livello': None, 'verificato': True,
                    'nota': 'provincia non presente nel servizio e-Distribuzione'}
        a = f[0]['attributes']
        lv = a.get('livellocriticita')
        et, spieg = SCALA.get(lv, ('ignoto', 'valore fuori scala'))
        return {'provincia': a.get('f4'), 'regione': a.get('f2'),
                'cod_pro': a.get('cod_pro'), 'livello': lv,
                'etichetta': et, 'significato': spieg, 'verificato': True,
                'granularita': 'PROVINCIALE (non la singola cabina primaria)',
                'fonte': 'e-Distribuzione Aree Critiche (feature service pubblico)',
                'scala_nota': 'ordinamento ricavato per evidenza dai dati, non da '
                              'documentazione ufficiale: crescente = peggio'}
    except Exception as e:
        return {'provincia': nome, 'livello': None, 'verificato': False,
                'nota': f'servizio non raggiunto ({type(e).__name__}): criticita NON verificata'}


def coda_da_export(path, prov=None, comune=None):
    """Legge un export Econnextion (xlsx/csv) e somma la coda per stato.

    L'export lo scarichi tu dalla dashboard: e' PowerBI, non ha API. Le colonne
    attese sono quelle dell'export standard (Regione, Provincia, Comune, Tipo
    Impianto, Fonte, Stato Connessione, Potenza (MW)).
    """
    righe = []
    if path.lower().endswith(('.xlsx', '.xlsm')):
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        it = ws.iter_rows(values_only=True)
        head = [str(h or '').strip().lower() for h in next(it)]
        for r in it:
            righe.append(dict(zip(head, r)))
    else:
        import csv
        with open(path, encoding='utf-8-sig') as fh:
            for r in csv.DictReader(fh, delimiter=';'):
                righe.append({k.strip().lower(): v for k, v in r.items()})

    def col(d, *names):
        for n in names:
            for k in d:
                if n in k:
                    return d[k]
        return None

    per_stato, per_fonte, per_tipo = {}, {}, {}
    tot = 0.0
    n = 0
    for d in righe:
        p = str(col(d, 'provincia') or '').strip().upper()
        c = str(col(d, 'comune') or '').strip().upper()
        if prov and p != prov.strip().upper():
            continue
        if comune and c != comune.strip().upper():
            continue
        try:
            mw = float(str(col(d, 'potenza') or 0).replace(',', '.'))
        except ValueError:
            continue
        st = str(col(d, 'stato') or 'ignoto').strip()
        fo = str(col(d, 'fonte') or 'ignota').strip()
        ti = str(col(d, 'tipo impianto', 'tipo') or 'ignoto').strip()
        per_tipo[ti] = per_tipo.get(ti, 0.0) + mw
        per_stato[st] = per_stato.get(st, 0.0) + mw
        per_fonte[fo] = per_fonte.get(fo, 0.0) + mw
        tot += mw
        n += 1

    # coda "pesata": una richiesta appena depositata non occupa la rete come un
    # progetto autorizzato. Senza pesi il totale sovrastima la congestione.
    pesata = 0.0
    for st, mw in per_stato.items():
        w = PESO_STATO.get(st.strip().lower(), 0.5)
        pesata += mw * w

    return {'file': os.path.basename(path), 'righe': n,
            'prov': prov, 'comune': comune,
            'righe_lette': len(righe),
            # Un filtro che non trova nulla in un file pieno non e' una coda vuota:
            # e' un filtro sbagliato (tipico: si passa la sigla 'BN' dove l'export
            # scrive 'BENEVENTO'). Dichiararlo evita di leggere "rete libera".
            'filtro_a_vuoto': bool(righe and n == 0),
            'mw_totali': round(tot, 1), 'mw_pesati': round(pesata, 1),
            'per_stato': {k: round(v, 1) for k, v in
                          sorted(per_stato.items(), key=lambda x: -x[1])},
            'per_fonte': {k: round(v, 1) for k, v in
                          sorted(per_fonte.items(), key=lambda x: -x[1])},
            'per_tipo': {k: round(v, 1) for k, v in
                         sorted(per_tipo.items(), key=lambda x: -x[1])},
            # un export di una sola tecnologia NON e' la coda del nodo: va detto,
            # altrimenti si legge "5.556 MW in coda" come se fosse tutto il fotovoltaico
            'copertura_parziale': (list(per_tipo)[0] if len(per_tipo) == 1 else None),
            'nota': 'export manuale dalla dashboard Terna Econnextion (PowerBI, '
                    'senza API pubblica): il dato vale alla data dell export'}


RE_RIGA = re.compile(
    r'^([A-Za-zÀ-ù\'\- ]+?)\s+([A-Z]{2})\s+(.+?)\s+(\d{2})\s+(x(?:\s+x)?)\s*$')


def inversioni_da_pdf(path, prov=None):
    """Sezioni AT/MT con inversione del flusso di energia (TICA art. 4.2 lett. c).

    E' il dato di saturazione alla granularita' che serve davvero: **la cabina
    primaria, con il suo nome**, non la provincia. Se una sezione esporta piu' di
    quanto consuma per oltre il 5% delle ore dell'anno, quella cabina e' gia'
    satura di generazione — e una nuova connessione li' e' onerosa o rifiutata.

    e-Distribuzione lo pubblica come PDF annuale (nessuna API): si scarica una
    volta e si legge. A differenza del livello provinciale, qui non c'e' niente
    da interpretare: e' un elenco.
    """
    import pdfplumber
    with pdfplumber.open(path) as pdf:
        txt = '\n'.join((p.extract_text() or '') for p in pdf.pages)
    out = []
    for l in txt.splitlines():
        m = RE_RIGA.match(l.strip())
        if not m:
            continue
        regione, pr, cab, sez, x = m.groups()
        if prov and pr.upper() != prov.strip().upper():
            continue
        out.append({'regione': regione.strip(), 'prov': pr, 'cabina': cab.strip(),
                    'sezione': sez, 'inv_1pct': True, 'inv_5pct': x.count('x') >= 2})
    return {'file': os.path.basename(path), 'prov': prov, 'sezioni': out,
            'cabine': sorted({r['cabina'] for r in out}),
            'cabine_5pct': sorted({r['cabina'] for r in out if r['inv_5pct']}),
            'nota': 'elenco ufficiale e-Distribuzione (TICA art. 4.2 lett. c): '
                    'inversione di flusso = la sezione esporta piu di quanto consuma'}


def cabina_critica(inv, nome_cabina):
    """La cabina di riferimento e' fra quelle sature?

    Confronto tollerante sui nomi: nel PDF sono abbreviati ("BENEVEN.NORD").
    """
    if not inv or not nome_cabina:
        return None
    n = re.sub(r'[^A-Z]', '', (nome_cabina or '').upper())
    for r in inv['sezioni']:
        c = re.sub(r'[^A-Z]', '', r['cabina'].upper())
        if n and (n.startswith(c[:8]) or c.startswith(n[:8])):
            return r
    return None


def valuta(crit, coda=None, mwp=None, inversioni=None, cabina=None):
    """Verdetto sul nodo: rischi e punti forti, mai un punteggio unico.

    Un numero solo nasconderebbe la differenza fra "non lo so" e "va bene",
    che qui e' la distinzione piu' importante.
    """
    rischi, punti, aperti = [], [], []

    lv = (crit or {}).get('livello')
    if not (crit or {}).get('verificato'):
        aperti.append('criticita di rete NON VERIFICATA (servizio non raggiunto): '
                      'da ricontrollare prima di qualunque impegno')
    elif lv is None:
        aperti.append(f"provincia non presente nel servizio e-Distribuzione: "
                      f"{crit.get('nota', '')}")
    elif lv >= 4:
        rischi.append(f"rete {crit['etichetta']} (livello {lv}/4): {crit['significato']} "
                      f"— verificare la connessione PRIMA di spendere su terra e visure")
    elif lv == 3:
        rischi.append(f"rete {crit['etichetta']} (livello {lv}/4): {crit['significato']}")
    else:
        punti.append(f"rete {crit['etichetta']} (livello {lv}/4): {crit['significato']}")
    if lv is not None:
        aperti.append('il dato e PROVINCIALE: la cabina primaria di riferimento va '
                      'verificata con il gestore (richiesta di preventivo o STMG)')

    if coda and coda.get('filtro_a_vuoto'):
        aperti.append(
            f"export letto ({coda['righe_lette']} righe) ma il filtro provincia "
            f"'{coda.get('prov')}' non ha trovato NESSUNA riga: la coda qui non e "
            f"vuota, e' non misurata (l export scrive il nome esteso, non la sigla)")
    elif coda:
        rap = None
        if mwp:
            base = mwp[1] if isinstance(mwp, (tuple, list)) else mwp
            rap = coda['mw_pesati'] / base if base else None
        cp = coda.get('copertura_parziale')
        s = (f"coda di connessione in provincia: {coda['mw_totali']} MW richiesti "
             f"({coda['mw_pesati']} MW pesati per stato di avanzamento)"
             + (f" — SOLO '{cp}'" if cp else ''))
        if cp:
            aperti.append(f"l export copre solo la tecnologia '{cp}': la coda complessiva "
                          f"del nodo e maggiore, servono anche gli altri export")
        if rap and rap > 100:
            rischi.append(s + f" — sono ~{rap:.0f} volte la taglia di questo progetto: "
                              f"la fila davanti e' il vero collo di bottiglia")
        elif rap and rap > 20:
            rischi.append(s + f" — ~{rap:.0f}x la taglia di questo progetto")
        else:
            punti.append(s)
        aperti.append('la coda e per PROVINCIA e alla data dell export, non per nodo: '
                      'due progetti nella stessa provincia possono insistere su cabine diverse')
    else:
        aperti.append('coda di connessione NON verificata: serve un export Econnextion '
                      '(dashboard PowerBI, nessuna API pubblica)')

    # il dato per CABINA batte quello provinciale: e' la granularita' vera
    if inversioni:
        if cabina:
            hit = cabina_critica(inversioni, cabina)
            if hit:
                g = '>=5%' if hit['inv_5pct'] else '>=1%'
                rischi.append(
                    f"cabina primaria {hit['cabina']} sez. {hit['sezione']}: inversione di "
                    f"flusso {g} delle ore/anno — la sezione esporta gia piu di quanto "
                    f"consuma, nuova generazione difficile da accettare")
            else:
                punti.append(f"la cabina {cabina} NON compare fra le sezioni con inversione "
                             f"di flusso in provincia")
        else:
            aperti.append(
                f"in provincia {len(inversioni['cabine'])} cabine hanno inversione di flusso "
                f"({', '.join(inversioni['cabine'][:6])}): identificare su quale si "
                f"connetterebbe l impianto")
    else:
        aperti.append('inversioni di flusso per cabina NON verificate: scaricare il PDF '
                      'annuale e-Distribuzione (TICA art. 4.2 lett. c)')

    return {'criticita': crit, 'coda': coda, 'inversioni': inversioni, 'rischi': rischi,
            'punti_forti': punti, 'da_verificare': aperti}


def print_valutazione(v):
    c = v.get('criticita') or {}
    print('\n=== CAPACITA DI RETE ===')
    if c.get('livello') is not None:
        print(f"  {c.get('provincia')}: criticita {c.get('etichetta')} "
              f"(livello {c['livello']}/4) — {c.get('granularita')}")
    for x in v['punti_forti']:
        print(f'  + {x}')
    for x in v['rischi']:
        print(f'  ! {x}')
    for x in v['da_verificare']:
        print(f'  ? {x}')
    if c.get('scala_nota'):
        print(f"  ~ {c['scala_nota']}")


def main():
    import argparse
    ap = argparse.ArgumentParser(description='Capacita e criticita del nodo di rete.')
    ap.add_argument('--prov', required=True, help='nome provincia, es. Benevento')
    ap.add_argument('--coda', default=None, help='export Econnextion (xlsx/csv)')
    ap.add_argument('--prov-coda', default=None, help='nome provincia come scritto nell export')
    ap.add_argument('--mwp', type=float, default=None, help='taglia del progetto in MWp')
    ap.add_argument('--inversioni', default=None,
                    help='PDF e-Distribuzione delle sezioni AT/MT con inversione di flusso')
    ap.add_argument('--cabina', default=None,
                    help='nome della cabina primaria di riferimento, es. PONTELANDOLFO')
    A = ap.parse_args()
    c = criticita_provincia(A.prov)
    q = coda_da_export(A.coda, prov=(A.prov_coda or A.prov)) if A.coda else None
    inv = inversioni_da_pdf(A.inversioni, prov=None) if A.inversioni else None
    if inv and c.get('cod_pro'):
        sigla = {62: 'BN'}.get(c['cod_pro'])
        if sigla:
            inv = inversioni_da_pdf(A.inversioni, prov=sigla)
    v = valuta(c, q, mwp=A.mwp, inversioni=inv, cabina=A.cabina)
    print_valutazione(v)
    if q:
        print('\n  coda per stato:')
        for k, mw in list(q['per_stato'].items())[:6]:
            print(f'     {k:<32s} {mw:>10.1f} MW')


if __name__ == '__main__':
    main()
